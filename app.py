import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from supabase import create_client, Client

# Sayfa ayarı
st.set_page_config(page_title="Bölge Dashboard", layout="wide", page_icon="🌍")

# ==================== SUPABASE BAĞLANTISI ====================
SUPABASE_URL = "https://tlcgcdiycgfxpxwzkwuf.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRsY2djZGl5Y2dmeHB4d3prd3VmIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjU2NDgwMjksImV4cCI6MjA4MTIyNDAyOX0.4GnWTvUmdLzqcP0v8MAqaNUQkYgk0S8qrw6nSPsz-t4"

@st.cache_resource
def get_supabase_client():
    return create_client(SUPABASE_URL, SUPABASE_KEY)

supabase: Client = get_supabase_client()

# ==================== GİRİŞ SİSTEMİ ====================
USERS = {
    "ziya": "Gm2025!",
    "sm1": "Sm12025!",
    "sm2": "Sm22025!",
    "sm3": "Sm32025!",
    "sm4": "Sm42025!",
    "sma": "Sma2025!",
}

def login():
    if "user" not in st.session_state:
        st.session_state.user = None
    
    if st.session_state.user is None:
        st.markdown("""
        <div style="max-width: 400px; margin: 100px auto; padding: 40px; 
                    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); 
                    border-radius: 15px; text-align: center;">
            <h1 style="color: white;">🌍 Bölge Dashboard</h1>
            <p style="color: #aaa;">Envanter Risk Analizi</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            st.markdown("### 🔐 Giriş Yap")
            username = st.text_input("Kullanıcı Adı", key="login_user")
            password = st.text_input("Şifre", type="password", key="login_pass")
            
            if st.button("Giriş", use_container_width=True):
                if username.lower() in USERS and USERS[username.lower()] == password:
                    st.session_state.user = username.lower()
                    st.rerun()
                else:
                    st.error("❌ Hatalı kullanıcı adı veya şifre")
        st.stop()

login()

# ==================== SUPABASE FONKSİYONLARI ====================

def get_available_periods_from_supabase():
    """Mevcut dönemleri al"""
    try:
        result = supabase.table('envanter_veri').select('envanter_donemi').execute()
        if result.data:
            periods = list(set([r['envanter_donemi'] for r in result.data if r['envanter_donemi']]))
            return sorted(periods, reverse=True)
    except:
        pass
    return []


def get_available_sms_from_supabase():
    """Mevcut Satış Müdürlerini al"""
    try:
        result = supabase.table('envanter_veri').select('satis_muduru').execute()
        if result.data:
            sms = list(set([r['satis_muduru'] for r in result.data if r['satis_muduru']]))
            return sorted(sms)
    except:
        pass
    return []


def get_data_from_supabase(satis_muduru=None, donemler=None):
    """Supabase'den veri çek ve DataFrame'e çevir - Pagination ile tüm veriyi al"""
    try:
        all_data = []
        batch_size = 1000
        offset = 0
        
        while True:
            query = supabase.table('envanter_veri').select('*')
            
            if satis_muduru:
                query = query.eq('satis_muduru', satis_muduru)
            
            if donemler and len(donemler) > 0:
                query = query.in_('envanter_donemi', donemler)
            
            query = query.range(offset, offset + batch_size - 1)
            result = query.execute()
            
            if not result.data or len(result.data) == 0:
                break
            
            all_data.extend(result.data)
            
            if len(result.data) < batch_size:
                break
            
            offset += batch_size
        
        if not all_data:
            return pd.DataFrame()
        
        df = pd.DataFrame(all_data)
        
        # Sütun isimlerini geri çevir
        reverse_mapping = {
            'magaza_kodu': 'Mağaza Kodu',
            'magaza_tanim': 'Mağaza Adı',
            'satis_muduru': 'Satış Müdürü',
            'bolge_sorumlusu': 'Bölge Sorumlusu',
            'depolama_kosulu_grubu': 'Depolama Koşulu Grubu',
            'depolama_kosulu': 'Depolama Koşulu',
            'envanter_donemi': 'Envanter Dönemi',
            'envanter_tarihi': 'Envanter Tarihi',
            'envanter_baslangic_tarihi': 'Envanter Başlangıç Tarihi',
            'urun_grubu_kodu': 'Ürün Grubu Kodu',
            'urun_grubu_tanimi': 'Ürün Grubu Tanımı',
            'mal_grubu_kodu': 'Mal Grubu Kodu',
            'mal_grubu_tanimi': 'Mal Grubu Tanımı',
            'malzeme_kodu': 'Malzeme Kodu',
            'malzeme_tanimi': 'Malzeme Adı',
            'satis_fiyati': 'Satış Fiyatı',
            'sayim_miktari': 'Sayım Miktarı',
            'sayim_tutari': 'Sayım Tutarı',
            'kaydi_miktar': 'Kaydi Miktar',
            'kaydi_tutar': 'Kaydi Tutar',
            'fark_miktari': 'Fark Miktarı',
            'fark_tutari': 'Fark Tutarı',
            'kismi_envanter_miktari': 'Kısmi Envanter Miktarı',
            'kismi_envanter_tutari': 'Kısmi Envanter Tutarı',
            'fire_miktari': 'Fire Miktarı',
            'fire_tutari': 'Fire Tutarı',
            'onceki_fark_miktari': 'Önceki Fark Miktarı',
            'onceki_fark_tutari': 'Önceki Fark Tutarı',
            'onceki_fire_miktari': 'Önceki Fire Miktarı',
            'onceki_fire_tutari': 'Önceki Fire Tutarı',
            'satis_miktari': 'Satış Miktarı',
            'satis_hasilati': 'Satış Tutarı',
            'iade_miktari': 'İade Miktarı',
            'iade_tutari': 'İade Tutarı',
            'iptal_fisteki_miktar': 'İptal Fişteki Miktar',
            'iptal_fis_tutari': 'İptal Fiş Tutarı',
            'iptal_gp_miktari': 'İptal GP Miktarı',
            'iptal_gp_tutari': 'İptal GP Tutarı',
            'iptal_satir_miktari': 'İptal Satır Miktarı',
            'iptal_satir_tutari': 'İptal Satır Tutarı',
        }
        
        df = df.rename(columns=reverse_mapping)
        
        return df
        
    except Exception as e:
        st.error(f"Supabase hatası: {str(e)}")
        return pd.DataFrame()


# ==================== ANA UYGULAMA ====================

# Çıkış butonu sağ üstte
col_title, col_user = st.columns([4, 1])
with col_title:
    st.title("🌍 Bölge Dashboard")
with col_user:
    st.markdown(f"👤 **{st.session_state.user.upper()}**")
    if st.button("🚪 Çıkış", key="logout_btn"):
        st.session_state.user = None
        st.rerun()

# CSS
st.markdown("""
<style>
    .risk-kritik { background-color: #ff4444; color: white; padding: 15px; border-radius: 8px; text-align: center; font-weight: bold; font-size: 1.2rem; }
    .risk-riskli { background-color: #ff8800; color: white; padding: 15px; border-radius: 8px; text-align: center; font-weight: bold; font-size: 1.2rem; }
    .risk-dikkat { background-color: #ffcc00; color: black; padding: 15px; border-radius: 8px; text-align: center; font-weight: bold; font-size: 1.2rem; }
    .risk-temiz { background-color: #00cc66; color: white; padding: 15px; border-radius: 8px; text-align: center; font-weight: bold; font-size: 1.2rem; }
    
    .magaza-card {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        border-radius: 10px;
        padding: 15px;
        margin: 5px;
        color: white;
        border-left: 4px solid #ff4444;
    }
    .magaza-card.riskli { border-left-color: #ff8800; }
    .magaza-card.dikkat { border-left-color: #ffcc00; }
    .magaza-card.temiz { border-left-color: #00cc66; }
    
    .metric-box {
        background: #f0f2f6;
        border-radius: 8px;
        padding: 10px;
        text-align: center;
    }
    
    @media (max-width: 768px) {
        .stMetric { font-size: 0.8rem; }
        div[data-testid="column"] { padding: 0.25rem !important; }
    }
</style>
""", unsafe_allow_html=True)

# 10 TL Ürün Kodları (209 adet)
KASA_AKTIVITESI_KODLARI = {
    '25006448', '12002256', '12002046', '22001972', '12003295', '22002759', '22002500', '11002886', '22002215', '22002214',
    '22002259', '22002349', '16002163', '22002717', '16001587', '13001073', '30000944', '18002488', '17003609', '22002296',
    '22002652', '24004136', '24004137', '12003073', '22002328', '24005228', '24006215', '24005232', '24005231', '24006214',
    '24006212', '16002332', '16002342', '23001397', '16002310', '24001063', '24004020', '13002613', '13002317', '13002506',
    '16002285', '16002219', '16002286', '16002218', '13000258', '13000257', '13000256', '13000260', '13002533', '22002611',
    '22002579', '13002559', '13000187', '13002904', '13000189', '13000190', '13002908', '13001872', '13001874', '30000838',
    '30000926', '22002605', '22002604', '22002603', '12003241', '16002194', '16001734', '25005580', '25000237', '25000049',
    '16002099', '23001367', '23001510', '23001177', '23001403', '23001278', '22002732', '22002576', '22002577', '25006483',
    '23001240', '16002317', '30000958', '30000956', '24005155', '24005154', '24005156', '24005157', '24005153', '22000280',
    '22002773', '22002774', '22002501', '22002225', '22000397', '22001395', '22000396', '16001859', '18002956', '17003542',
    '16002338', '16002339', '16002341', '16002009', '16000856', '22002715', '16002235', '24006067', '24006069', '24006068',
    '24006066', '22002686', '22002687', '22002688', '16002220', '24005291', '24005290', '24006078', '24006084', '24005288',
    '24006082', '24006079', '24005289', '24006085', '22002763', '22002762', '22001032', '18003049', '24006126', '24004420',
    '24005183', '24005649', '24005650', '14002481', '13002315', '22001229', '13002478', '30000880', '24005798', '24005796',
    '24005799', '24005797', '24005795', '24006159', '24003492', '24006171', '24006170', '24006174', '24006172', '24006173',
    '22002640', '22002553', '22002764', '22002223', '22002679', '22002221', '22002224', '22002572', '27002662', '24005441',
    '24005897', '24005898', '24005900', '24006081', '24006080', '16002087', '22002282', '22002283', '24005893', '24005894',
    '23001198', '23001439', '23001195', '23001199', '23000843', '23000034', '23001445', '23001444', '23001443', '23001522',
    '24004381', '24005184', '23001534', '23001533', '18001591', '27002676', '27002677', '16001956', '24003287', '24000005',
    '24002194', '24002192', '24002764', '24003872', '16001983', '18002969', '27001340', '27001148', '27001563', '24004354',
    '24004196', '24004115', '14002424', '24003641', '24004972', '13001481', '24003327', '24000004', '23000122',
}


def analyze_inventory(df):
    """Veriyi analiz için hazırla"""
    df = df.copy()
    
    col_mapping = {
        'Mağaza Tanım': 'Mağaza Adı',
        'Malzeme Tanımı': 'Malzeme Adı',
        'Mal Grubu Tanımı': 'Ürün Grubu',
        'Satış Hasılatı': 'Satış Tutarı',
        'Satış Fiyatı': 'Birim Fiyat',
    }
    
    for old_col, new_col in col_mapping.items():
        if old_col in df.columns:
            df[new_col] = df[old_col]
    
    numeric_cols = ['Fark Miktarı', 'Fark Tutarı', 'Kısmi Envanter Miktarı', 'Kısmi Envanter Tutarı',
                    'Önceki Fark Miktarı', 'Önceki Fark Tutarı', 'Fire Miktarı', 'Fire Tutarı',
                    'Satış Miktarı', 'Satış Tutarı', 'Önceki Fire Miktarı', 'Önceki Fire Tutarı', 'Birim Fiyat']
    
    for col in numeric_cols:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # Toplam hesaplamaları
    df['Kısmi Envanter Tutarı'] = df.get('Kısmi Envanter Tutarı', pd.Series([0]*len(df))).fillna(0)
    df['Önceki Fark Tutarı'] = df.get('Önceki Fark Tutarı', pd.Series([0]*len(df))).fillna(0)
    df['TOPLAM_FARK'] = df['Fark Tutarı'] + df['Kısmi Envanter Tutarı'] + df['Önceki Fark Tutarı']
    
    return df


def detect_internal_theft(df):
    """İç hırsızlık tespiti - Satış Fiyatı ≥100TL ve açık"""
    results = []
    for idx, row in df.iterrows():
        fiyat = row.get('Birim Fiyat', 0) or 0
        fark = row['Fark Miktarı']
        kismi = row['Kısmi Envanter Miktarı']
        onceki = row['Önceki Fark Miktarı']
        toplam = fark + kismi + onceki
        
        if fiyat >= 100 and toplam < 0:
            results.append(row)
    return pd.DataFrame(results)


def detect_chronic_shortage(df):
    """Kronik açık - Her iki dönemde de Fark < 0 ve dengelenmemiş"""
    results = []
    for idx, row in df.iterrows():
        onceki = row.get('Önceki Fark Miktarı', 0) or 0
        bu_donem = row['Fark Miktarı']
        
        if onceki < 0 and bu_donem < 0:
            if abs(onceki + bu_donem) > 0.01:  # Dengelenmemiş
                results.append(row)
    return pd.DataFrame(results)


def detect_cigarette_shortage(df):
    """
    Sigara açığı tespiti - TOPLAM BAZLI
    Tüm sigaraların (Fark + Kısmi + Önceki) toplamı < 0 ise açık var
    Dönen değer: Açık varsa 1, yoksa 0 (veya açık miktarı)
    """
    toplam_fark = 0
    toplam_kismi = 0
    toplam_onceki = 0
    sigara_var = False
    
    for idx, row in df.iterrows():
        urun_grubu = str(row.get('Ürün Grubu', '')).upper()
        mal_grubu = str(row.get('Mal Grubu Tanımı', '')).upper()
        malzeme = str(row.get('Malzeme Adı', '')).upper()
        
        is_cigarette = any(x in urun_grubu or x in mal_grubu or x in malzeme 
                          for x in ['SİGARA', 'SIGARA', 'TOBACCO', 'TÜTÜN'])
        
        if is_cigarette:
            sigara_var = True
            fark = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
            kismi = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
            onceki = row['Önceki Fark Miktarı'] if pd.notna(row['Önceki Fark Miktarı']) else 0
            
            toplam_fark += fark
            toplam_kismi += kismi
            toplam_onceki += onceki
    
    if not sigara_var:
        return pd.DataFrame()
    
    net_toplam = toplam_fark + toplam_kismi + toplam_onceki
    
    # Eğer net toplam < 0 ise açık var, 1 satırlık DataFrame döndür
    if net_toplam < 0:
        return pd.DataFrame([{
            'Açık Miktarı': abs(net_toplam),
            'Fark Toplam': toplam_fark,
            'Kısmi Toplam': toplam_kismi,
            'Önceki Toplam': toplam_onceki,
            'Net Toplam': net_toplam
        }])
    
    return pd.DataFrame()


def check_10tl_products(df):
    """10 TL ürünleri kontrolü - FORMÜL: Fark + Kısmi (Önceki dahil değil)"""
    toplam_adet = 0
    toplam_tutar = 0
    
    for idx, row in df.iterrows():
        kod_str = str(row.get('Malzeme Kodu', '')).replace('.0', '').strip()
        
        if kod_str in KASA_AKTIVITESI_KODLARI:
            fark = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
            kismi = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
            toplam = fark + kismi  # Önceki dahil değil!
            
            fark_tutari = row.get('Fark Tutarı', 0) or 0
            kismi_tutari = row.get('Kısmi Envanter Tutarı', 0) or 0
            
            toplam_adet += toplam
            toplam_tutar += fark_tutari + kismi_tutari  # Önceki dahil değil!
    
    return {'adet': toplam_adet, 'tutar': toplam_tutar}


def calculate_risk_score(toplam_oran, sigara_count, ic_hirsizlik_count, kronik_count, kasa_adet, bolge_ort):
    """
    Risk puanı hesaplama (0-100)
    Ağırlıklar:
    - Toplam Oran: %30
    - Sigara Açığı: %30
    - İç Hırsızlık: %30
    - Kronik Açık: %5
    - 10TL Ürünleri: %5
    """
    puan = 0
    
    # Toplam Oran (30 puan) - Bölge ortalamasına göre
    if bolge_ort['toplam_oran'] > 0:
        oran_ratio = toplam_oran / bolge_ort['toplam_oran']
        oran_puan = min(30, oran_ratio * 15)  # 2x ortalama = 30 puan
    else:
        oran_puan = min(30, toplam_oran * 20)
    puan += oran_puan
    
    # Sigara Açığı (30 puan) - Her sigara kritik
    if sigara_count > 10:
        sigara_puan = 30
    elif sigara_count > 5:
        sigara_puan = 25
    elif sigara_count > 0:
        sigara_puan = sigara_count * 4
    else:
        sigara_puan = 0
    puan += sigara_puan
    
    # İç Hırsızlık (30 puan) - Bölge ortalamasına göre
    if bolge_ort['ic_hirsizlik'] > 0:
        ic_ratio = ic_hirsizlik_count / bolge_ort['ic_hirsizlik']
        ic_puan = min(30, ic_ratio * 15)
    else:
        ic_puan = min(30, ic_hirsizlik_count * 0.5)
    puan += ic_puan
    
    # Kronik Açık (5 puan)
    if bolge_ort['kronik'] > 0:
        kronik_ratio = kronik_count / bolge_ort['kronik']
        kronik_puan = min(5, kronik_ratio * 2.5)
    else:
        kronik_puan = min(5, kronik_count * 0.05)
    puan += kronik_puan
    
    # 10TL Ürünleri (5 puan) - Fazla = şüpheli
    if kasa_adet > 20:
        kasa_puan = 5
    elif kasa_adet > 10:
        kasa_puan = 3
    elif kasa_adet > 0:
        kasa_puan = 1
    else:
        kasa_puan = 0
    puan += kasa_puan
    
    return min(100, max(0, puan))


def get_risk_level(puan):
    """Risk seviyesi belirleme"""
    if puan >= 60:
        return "🔴 KRİTİK", "kritik"
    elif puan >= 40:
        return "🟠 RİSKLİ", "riskli"
    elif puan >= 20:
        return "🟡 DİKKAT", "dikkat"
    else:
        return "🟢 TEMİZ", "temiz"


def analyze_store(df_store):
    """Tek mağaza analizi"""
    satis = df_store['Satış Tutarı'].sum()
    
    # Fark = Fark Tutarı + Kısmi Envanter Tutarı
    fark_tutari = df_store['Fark Tutarı'].fillna(0).sum()
    kismi_tutari = df_store['Kısmi Envanter Tutarı'].fillna(0).sum()
    fark = fark_tutari + kismi_tutari
    
    # Fire = Fire Tutarı
    fire = df_store['Fire Tutarı'].fillna(0).sum()
    
    # Toplam Açık = Fark + Fire (yani Fark Tutarı + Kısmi + Fire)
    toplam_acik = fark + fire
    
    # Oranlar
    fark_oran = abs(fark) / satis * 100 if satis > 0 else 0
    fire_oran = abs(fire) / satis * 100 if satis > 0 else 0
    toplam_oran = abs(toplam_acik) / satis * 100 if satis > 0 else 0
    
    # Gün hesabı
    gun_sayisi = 1
    try:
        if 'Envanter Tarihi' in df_store.columns and 'Envanter Başlangıç Tarihi' in df_store.columns:
            env_tarihi = pd.to_datetime(df_store['Envanter Tarihi'].iloc[0])
            env_baslangic = pd.to_datetime(df_store['Envanter Başlangıç Tarihi'].iloc[0])
            gun_sayisi = (env_tarihi - env_baslangic).days
            if gun_sayisi <= 0:
                gun_sayisi = 1
    except:
        gun_sayisi = 1
    
    gunluk_fark = fark / gun_sayisi
    gunluk_fire = fire / gun_sayisi
    
    internal_df = detect_internal_theft(df_store)
    chronic_df = detect_chronic_shortage(df_store)
    cigarette_df = detect_cigarette_shortage(df_store)
    kasa_result = check_10tl_products(df_store)
    
    return {
        'satis': satis,
        'fark': fark,
        'fire': fire,
        'toplam_acik': toplam_acik,
        'fark_oran': fark_oran,
        'fire_oran': fire_oran,
        'toplam_oran': toplam_oran,
        'gun_sayisi': gun_sayisi,
        'gunluk_fark': gunluk_fark,
        'gunluk_fire': gunluk_fire,
        'ic_hirsizlik': len(internal_df),
        'kronik': len(chronic_df),
        'sigara': int(cigarette_df['Açık Miktarı'].iloc[0]) if len(cigarette_df) > 0 else 0,
        'kasa_adet': kasa_result['adet'],
        'kasa_tutar': kasa_result['tutar']
    }


def analyze_all_stores(df):
    """Tüm mağazaları analiz et"""
    magazalar = df['Mağaza Kodu'].dropna().unique().tolist()
    results = []
    
    # Önce tüm mağazaları analiz et
    store_data = {}
    for mag in magazalar:
        df_mag = df[df['Mağaza Kodu'] == mag].copy()
        if len(df_mag) == 0:
            continue
        
        mag_adi = df_mag['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df_mag.columns else ''
        sm = df_mag['Satış Müdürü'].iloc[0] if 'Satış Müdürü' in df_mag.columns else ''
        bs = df_mag['Bölge Sorumlusu'].iloc[0] if 'Bölge Sorumlusu' in df_mag.columns else ''
        
        metrics = analyze_store(df_mag)
        store_data[mag] = {
            'kod': mag,
            'adi': mag_adi,
            'sm': sm,
            'bs': bs,
            **metrics
        }
    
    # Bölge ortalamaları
    if len(store_data) > 0:
        bolge_ort = {
            'toplam_oran': np.mean([s['toplam_oran'] for s in store_data.values()]),
            'ic_hirsizlik': np.mean([s['ic_hirsizlik'] for s in store_data.values()]),
            'kronik': np.mean([s['kronik'] for s in store_data.values()]),
            'sigara': np.mean([s['sigara'] for s in store_data.values()]),
        }
    else:
        bolge_ort = {'toplam_oran': 1, 'ic_hirsizlik': 1, 'kronik': 1, 'sigara': 1}
    
    # Risk puanları hesapla
    for mag, data in store_data.items():
        risk_puan = calculate_risk_score(
            data['toplam_oran'],
            data['sigara'],
            data['ic_hirsizlik'],
            data['kronik'],
            data['kasa_adet'],
            bolge_ort
        )
        risk_seviye, risk_class = get_risk_level(risk_puan)
        
        # Risk nedenleri
        nedenler = []
        if data['sigara'] > 0:
            nedenler.append(f"🚬 Sigara:{data['sigara']}")
        if data['toplam_oran'] > bolge_ort['toplam_oran'] * 1.5:
            nedenler.append(f"📉 Toplam:%{data['toplam_oran']:.1f}")
        if data['ic_hirsizlik'] > bolge_ort['ic_hirsizlik'] * 1.5:
            nedenler.append(f"🔒 İç Hırs:{data['ic_hirsizlik']}")
        if data['kasa_adet'] > 10:
            nedenler.append(f"💰 10TL:+{data['kasa_adet']:.0f}/{data['kasa_tutar']:,.0f}₺")
        
        results.append({
            'Mağaza Kodu': mag,
            'Mağaza Adı': data['adi'],
            'SM': data['sm'],
            'BS': data['bs'],
            'Satış': data['satis'],
            'Fark': data['fark'],
            'Fire': data['fire'],
            'Toplam Açık': data['toplam_acik'],
            'Fark %': data['fark_oran'],
            'Fire %': data['fire_oran'],
            'Toplam %': data['toplam_oran'],
            'İç Hırs.': data['ic_hirsizlik'],
            'Kronik': data['kronik'],
            'Sigara': data['sigara'],
            '10TL Adet': data['kasa_adet'],
            '10TL Tutar': data['kasa_tutar'],
            'Gün': data['gun_sayisi'],
            'Günlük Fark': data['gunluk_fark'],
            'Günlük Fire': data['gunluk_fire'],
            'Risk Puan': risk_puan,
            'Risk': risk_seviye,
            'Risk Class': risk_class,
            'Nedenler': " | ".join(nedenler) if nedenler else "-"
        })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        result_df = result_df.sort_values('Risk Puan', ascending=False)
    
    return result_df, bolge_ort


def aggregate_by_group(store_df, group_col):
    """SM veya BS bazında gruplama"""
    if group_col not in store_df.columns:
        return pd.DataFrame()
    
    grouped = store_df.groupby(group_col).agg({
        'Mağaza Kodu': 'count',
        'Satış': 'sum',
        'Fark': 'sum',
        'Fire': 'sum',
        'Toplam Açık': 'sum',
        'İç Hırs.': 'sum',
        'Kronik': 'sum',
        'Sigara': 'sum',
        '10TL Adet': 'sum',
        'Gün': 'sum',
        'Risk Puan': 'mean'
    }).reset_index()
    
    grouped.columns = [group_col, 'Mağaza Sayısı', 'Satış', 'Fark', 'Fire', 'Toplam Açık',
                       'İç Hırs.', 'Kronik', 'Sigara', '10TL Adet', 'Toplam Gün', 'Ort. Risk']
    
    # Oranlar
    grouped['Fark %'] = abs(grouped['Fark']) / grouped['Satış'] * 100
    grouped['Fark %'] = grouped['Fark %'].fillna(0)
    
    grouped['Fire %'] = abs(grouped['Fire']) / grouped['Satış'] * 100
    grouped['Fire %'] = grouped['Fire %'].fillna(0)
    
    grouped['Toplam %'] = abs(grouped['Toplam Açık']) / grouped['Satış'] * 100
    grouped['Toplam %'] = grouped['Toplam %'].fillna(0)
    
    # Günlük fark ve fire
    grouped['Günlük Fark'] = grouped['Fark'] / grouped['Toplam Gün']
    grouped['Günlük Fark'] = grouped['Günlük Fark'].fillna(0)
    grouped['Günlük Fire'] = grouped['Fire'] / grouped['Toplam Gün']
    grouped['Günlük Fire'] = grouped['Günlük Fire'].fillna(0)
    
    # Risk seviyesi
    grouped['Risk'] = grouped['Ort. Risk'].apply(lambda x: get_risk_level(x)[0])
    
    # Kritik mağaza sayısı
    for idx, row in grouped.iterrows():
        grup_magazalar = store_df[store_df[group_col] == row[group_col]]
        kritik_count = len(grup_magazalar[grup_magazalar['Risk'].str.contains('KRİTİK')])
        grouped.at[idx, 'Kritik Mağaza'] = kritik_count
    
    grouped = grouped.sort_values('Ort. Risk', ascending=False)
    
    return grouped


def create_store_report(store_row, params, df_all=None):
    """Tek mağaza için detaylı Excel raporu - SM App ile birebir aynı format"""
    
    wb = Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    red_fill = PatternFill('solid', fgColor='FF4444')
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    mag_kod = store_row['Mağaza Kodu']
    mag_adi = store_row['Mağaza Adı']
    
    # Mağaza verisi yoksa sadece özet döndür
    if df_all is None:
        ws = wb.active
        ws.title = "ÖZET"
        ws['A1'] = f"MAĞAZA: {mag_kod} - {mag_adi}"
        ws['A1'].font = title_font
        ws['A2'] = "Detaylı veri yüklenemedi"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    df_mag = df_all[df_all['Mağaza Kodu'] == mag_kod].copy()
    
    if len(df_mag) == 0:
        ws = wb.active
        ws.title = "ÖZET"
        ws['A1'] = f"MAĞAZA: {mag_kod} - {mag_adi}"
        ws['A1'].font = title_font
        ws['A2'] = "Mağaza verisi bulunamadı"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    # ==================== ANALİZLER ====================
    
    # Temel metrikler
    toplam_satis = df_mag['Satış Tutarı'].sum()
    fark_tutari = df_mag['Fark Tutarı'].fillna(0).sum()
    kismi_tutari = df_mag['Kısmi Envanter Tutarı'].fillna(0).sum()
    fire_tutari = df_mag['Fire Tutarı'].fillna(0).sum()
    fark = fark_tutari + kismi_tutari
    toplam_acik = fark + fire_tutari
    fark_oran = abs(fark) / toplam_satis * 100 if toplam_satis > 0 else 0
    fire_oran = abs(fire_tutari) / toplam_satis * 100 if toplam_satis > 0 else 0
    toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
    
    # 1. İÇ HIRSIZLIK (Satış Fiyatı >= 100 TL, Fark <= 0)
    internal_results = []
    for idx, row in df_mag.iterrows():
        satis_fiyati = row.get('Satış Fiyatı', 0) or 0
        fark_m = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
        kismi_m = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
        onceki_m = row['Önceki Fark Miktarı'] if pd.notna(row['Önceki Fark Miktarı']) else 0
        toplam_m = fark_m + kismi_m + onceki_m
        iptal = row.get('İptal Satır Sayısı', 0) or 0
        
        if satis_fiyati >= 100 and toplam_m <= 0:
            # Risk ve durum hesapla
            if iptal > 0 and abs(toplam_m) == iptal:
                durum = "TAM EŞİT"
                risk = "ÇOK YÜKSEK"
            elif iptal > 0 and abs(abs(toplam_m) - iptal) <= 2:
                durum = "YAKIN (±2)"
                risk = "YÜKSEK"
            elif iptal > 0:
                durum = "İPTAL VAR"
                risk = "ORTA"
            else:
                durum = "İPTAL YOK"
                risk = "DÜŞÜK"
            
            internal_results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Ürün Grubu': row.get('Ürün Grubu', ''),
                'Satış Fiyatı': satis_fiyati,
                'Fark Miktarı': fark_m,
                'Kısmi Env.': kismi_m,
                'Önceki Fark': onceki_m,
                'TOPLAM': toplam_m,
                'İptal Satır': iptal,
                'Fark': toplam_m - (-iptal) if iptal > 0 else toplam_m,
                'Durum': durum,
                'Fark Tutarı (TL)': row.get('Fark Tutarı', 0),
                'Risk': risk
            })
    
    internal_df = pd.DataFrame(internal_results)
    if len(internal_df) > 0:
        internal_df = internal_df.sort_values('Fark Tutarı (TL)')
    
    # 2. KRONİK AÇIK (Bu dönem ve önceki dönem fark < 0)
    chronic_results = []
    for idx, row in df_mag.iterrows():
        fark_m = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
        onceki_m = row['Önceki Fark Miktarı'] if pd.notna(row['Önceki Fark Miktarı']) else 0
        
        if fark_m < 0 and onceki_m < 0:
            fark_t = row.get('Fark Tutarı', 0) or 0
            satis_fiyati = row.get('Satış Fiyatı', 0) or 0
            onceki_t = onceki_m * satis_fiyati if satis_fiyati else 0
            
            chronic_results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Ürün Grubu': row.get('Ürün Grubu', ''),
                'Bu Dönem Fark': fark_m,
                'Bu Dönem Tutar': fark_t,
                'Önceki Fark': onceki_m,
                'Önceki Tutar': onceki_t,
                'Toplam Tutar': fark_t + onceki_t
            })
    
    chronic_df = pd.DataFrame(chronic_results)
    if len(chronic_df) > 0:
        chronic_df = chronic_df.sort_values('Toplam Tutar')
    
    # 3. KRONİK FİRE (Bu dönem ve önceki dönem fire < 0)
    chronic_fire_results = []
    for idx, row in df_mag.iterrows():
        fire_m = row['Fire Miktarı'] if pd.notna(row['Fire Miktarı']) else 0
        onceki_fire = row.get('Önceki Fire Miktarı', 0) or 0
        
        if fire_m < 0 and onceki_fire < 0:
            fire_t = row.get('Fire Tutarı', 0) or 0
            satis_fiyati = row.get('Satış Fiyatı', 0) or 0
            onceki_fire_t = onceki_fire * satis_fiyati if satis_fiyati else 0
            
            chronic_fire_results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Ürün Grubu': row.get('Ürün Grubu', ''),
                'Bu Dönem Fire': fire_m,
                'Bu Dönem Fire Tutarı': fire_t,
                'Önceki Fire': onceki_fire,
                'Önceki Fire Tutarı': onceki_fire_t,
                'Toplam Fire Tutarı': fire_t + onceki_fire_t
            })
    
    chronic_fire_df = pd.DataFrame(chronic_fire_results)
    if len(chronic_fire_df) > 0:
        chronic_fire_df = chronic_fire_df.sort_values('Toplam Fire Tutarı')
    
    # 4. SİGARA AÇIĞI
    sigara_results = []
    for idx, row in df_mag.iterrows():
        urun_grubu = str(row.get('Ürün Grubu', '')).upper()
        mal_grubu = str(row.get('Mal Grubu Tanımı', '')).upper()
        malzeme = str(row.get('Malzeme Adı', '')).upper()
        
        is_cigarette = any(x in urun_grubu or x in mal_grubu or x in malzeme 
                          for x in ['SİGARA', 'SIGARA', 'TOBACCO', 'TÜTÜN', 'MAKARON'])
        
        if is_cigarette:
            fark_m = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
            kismi_m = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
            onceki_m = row['Önceki Fark Miktarı'] if pd.notna(row['Önceki Fark Miktarı']) else 0
            urun_toplam = fark_m + kismi_m + onceki_m
            
            if fark_m != 0 or kismi_m != 0 or onceki_m != 0:
                sigara_results.append({
                    'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                    'Malzeme Adı': row.get('Malzeme Adı', ''),
                    'Fark': fark_m,
                    'Kısmi': kismi_m,
                    'Önceki': onceki_m,
                    'Ürün Toplam': urun_toplam,
                    'Risk': 'SİGARA'
                })
    
    cigarette_df = pd.DataFrame(sigara_results)
    if len(cigarette_df) > 0:
        cigarette_df = cigarette_df.sort_values('Ürün Toplam')
    
    # 5. AİLE ANALİZİ (Benzer ürün grupları)
    family_results = []
    df_mag['_ilk2'] = df_mag['Malzeme Adı'].fillna('').str.split().str[:2].str.join(' ')
    df_mag['_marka'] = df_mag['Malzeme Adı'].fillna('').str.split().str[-1]
    
    grouped = df_mag.groupby(['Mal Grubu Tanımı', '_ilk2', '_marka'])
    for (mal_grubu, ilk2, marka), group in grouped:
        if len(group) >= 2:
            toplam_fark = group['Fark Miktarı'].fillna(0).sum()
            toplam_kismi = group['Kısmi Envanter Miktarı'].fillna(0).sum()
            toplam_onceki = group['Önceki Fark Miktarı'].fillna(0).sum()
            aile_toplam = toplam_fark + toplam_kismi + toplam_onceki
            
            if aile_toplam < 0:
                urunler = []
                for _, r in group.iterrows():
                    urun_adi = str(r.get('Malzeme Adı', ''))[:25]
                    urun_fark = r['Fark Miktarı'] if pd.notna(r['Fark Miktarı']) else 0
                    urunler.append(f"{urun_adi}({urun_fark})")
                
                family_results.append({
                    'Mal Grubu': mal_grubu,
                    'İlk 2 Kelime': ilk2,
                    'Marka': marka,
                    'Ürün Sayısı': len(group),
                    'Toplam Fark': toplam_fark,
                    'Toplam Kısmi': toplam_kismi,
                    'Toplam Önceki': toplam_onceki,
                    'AİLE TOPLAMI': aile_toplam,
                    'Sonuç': 'AİLEDE NET AÇIK VAR',
                    'Risk': 'ORTA',
                    'Ürünler': ' | '.join(urunler[:5])
                })
    
    family_df = pd.DataFrame(family_results)
    if len(family_df) > 0:
        family_df = family_df.sort_values('AİLE TOPLAMI')
    
    # 6. FİRE MANİPÜLASYONU (Fark > 0, Fire < 0)
    fire_manip_results = []
    for idx, row in df_mag.iterrows():
        fark_m = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
        kismi_m = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
        fire_m = row['Fire Miktarı'] if pd.notna(row['Fire Miktarı']) else 0
        
        fark_kismi = fark_m + kismi_m
        
        if fark_kismi > 0 and fire_m < 0:
            fire_manip_results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Ürün Grubu': row.get('Ürün Grubu', ''),
                'Fark Miktarı': fark_m,
                'Kısmi Env.': kismi_m,
                'Önceki Fark': row.get('Önceki Fark Miktarı', 0),
                'Fark + Kısmi': fark_kismi,
                'Fire Miktarı': fire_m,
                'Fire Tutarı': row.get('Fire Tutarı', 0),
                'Sonuç': 'FAZLA FİRE GİRİLMİŞ'
            })
    
    fire_manip_df = pd.DataFrame(fire_manip_results)
    if len(fire_manip_df) > 0:
        fire_manip_df = fire_manip_df.sort_values('Fire Tutarı')
    
    # 7. KASA AKTİVİTESİ (10TL Ürünleri)
    kasa_results = []
    for idx, row in df_mag.iterrows():
        kod_str = str(row.get('Malzeme Kodu', '')).replace('.0', '').strip()
        
        if kod_str in KASA_AKTIVITESI_KODLARI:
            fark_m = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
            kismi_m = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
            toplam_m = fark_m + kismi_m
            
            fark_t = row.get('Fark Tutarı', 0) or 0
            kismi_t = row.get('Kısmi Envanter Tutarı', 0) or 0
            toplam_t = fark_t + kismi_t
            
            if toplam_m != 0:
                kasa_results.append({
                    'Malzeme Kodu': kod_str,
                    'Malzeme Adı': row.get('Malzeme Adı', ''),
                    'Fark': fark_m,
                    'Kısmi': kismi_m,
                    'TOPLAM': toplam_m,
                    'Tutar': toplam_t,
                    'Durum': 'FAZLA (+)' if toplam_m > 0 else 'AÇIK (-)'
                })
    
    kasa_df = pd.DataFrame(kasa_results)
    if len(kasa_df) > 0:
        kasa_df = kasa_df.sort_values('TOPLAM', ascending=False)
    
    # 8. EN RİSKLİ 20
    top20_results = []
    df_sorted = df_mag.nsmallest(20, 'Fark Tutarı')
    for idx, row in df_sorted.iterrows():
        fark_m = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
        kismi_m = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
        onceki_m = row['Önceki Fark Miktarı'] if pd.notna(row['Önceki Fark Miktarı']) else 0
        toplam_m = fark_m + kismi_m + onceki_m
        iptal = row.get('İptal Satır Sayısı', 0) or 0
        fire_m = row.get('Fire Miktarı', 0) or 0
        fire_t = row.get('Fire Tutarı', 0) or 0
        fark_t = row.get('Fark Tutarı', 0) or 0
        satis_fiyati = row.get('Satış Fiyatı', 0) or 0
        
        # Risk türü belirleme
        if satis_fiyati >= 100 and toplam_m < 0:
            risk_turu = "İÇ HIRSIZLIK"
            aksiyon = "Kasa kamera incelemesi"
        elif fark_m < 0 and onceki_m < 0:
            risk_turu = "KRONİK AÇIK"
            aksiyon = "Raf kontrolü, Sayım eğitimi"
        elif fark_m > 0 and fire_m < 0:
            risk_turu = "OPERASYONEL"
            aksiyon = "Fire kayıt kontrolü"
        else:
            risk_turu = "DIŞ HIRSIZLIK/SAYIM"
            aksiyon = "Sayım ve kod kontrolü"
        
        top20_results.append({
            'Sıra': len(top20_results) + 1,
            'Malzeme Kodu': row.get('Malzeme Kodu', ''),
            'Malzeme Adı': row.get('Malzeme Adı', ''),
            'Fark Mik.': fark_m,
            'Kısmi': kismi_m,
            'Önceki': onceki_m,
            'TOPLAM': toplam_m,
            'İptal': iptal,
            'Fire': fire_m,
            'Fire Tutarı': fire_t,
            'Fark Tutarı': fark_t,
            'Risk Türü': risk_turu,
            'Aksiyon': aksiyon
        })
    
    top20_df = pd.DataFrame(top20_results)
    
    # ==================== EXCEL OLUŞTURMA ====================
    
    # ===== ÖZET SAYFASI =====
    ws = wb.active
    ws.title = "ÖZET"
    
    ws['A1'] = f"MAĞAZA: {mag_kod} - {mag_adi}"
    ws['A1'].font = title_font
    ws['A2'] = f"Dönem: {params.get('donem', '')} | Tarih: {params.get('tarih', '')}"
    
    ws['A4'] = "GENEL METRIKLER"
    ws['A4'].font = subtitle_font
    
    metrics = [
        ('Toplam Ürün', len(df_mag)),
        ('Açık Veren Ürün', len(df_mag[df_mag['Fark Miktarı'] < 0])),
        ('Toplam Satış', f"{toplam_satis:,.0f} TL"),
        ('Fark (Fark+Kısmi)', f"{fark:,.0f} TL"),
        ('Fire', f"{fire_tutari:,.0f} TL"),
        ('Toplam Açık', f"{toplam_acik:,.0f} TL"),
        ('Fark Oranı', f"%{fark_oran:.2f}"),
        ('Fire Oranı', f"%{fire_oran:.2f}"),
        ('Toplam Oran', f"%{toplam_oran:.2f}"),
    ]
    
    for i, (label, value) in enumerate(metrics, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
    
    ws['A15'] = "RİSK DAĞILIMI"
    ws['A15'].font = subtitle_font
    
    risks = [
        ('İç Hırsızlık (≥100TL)', len(internal_df)),
        ('Kronik Açık', len(chronic_df)),
        ('Kronik Fire', len(chronic_fire_df)),
        ('Sigara Açığı', len(cigarette_df)),
        ('Fire Manipülasyonu', len(fire_manip_df)),
    ]
    
    for i, (label, value) in enumerate(risks, start=16):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        if 'Sigara' in label and value > 0:
            ws[f'B{i}'].fill = red_fill
            ws[f'B{i}'].font = Font(bold=True, color='FFFFFF')
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    # ===== EN RİSKLİ 20 =====
    if len(top20_df) > 0:
        ws2 = wb.create_sheet("EN RİSKLİ 20")
        for col, h in enumerate(top20_df.columns, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for r_idx, row in enumerate(top20_df.values, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
    
    # ===== KRONİK AÇIK =====
    if len(chronic_df) > 0:
        ws3 = wb.create_sheet("KRONİK AÇIK")
        for col, h in enumerate(chronic_df.columns, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(chronic_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws3.cell(row=r_idx, column=c_idx, value=val)
    
    # ===== KRONİK FİRE =====
    if len(chronic_fire_df) > 0:
        ws4 = wb.create_sheet("KRONİK FİRE")
        for col, h in enumerate(chronic_fire_df.columns, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(chronic_fire_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws4.cell(row=r_idx, column=c_idx, value=val)
    
    # ===== SİGARA AÇIĞI =====
    ws5 = wb.create_sheet("SİGARA AÇIĞI")
    ws5['A1'] = "⚠️ SİGARA AÇIĞI - YÜKSEK RİSK"
    ws5['A1'].font = Font(bold=True, size=14, color='FF0000')
    
    if len(cigarette_df) > 0:
        for col, h in enumerate(cigarette_df.columns, 1):
            cell = ws5.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = red_fill
        
        for r_idx, row in enumerate(cigarette_df.values, 4):
            for c_idx, val in enumerate(row, 1):
                ws5.cell(row=r_idx, column=c_idx, value=val)
    
    # ===== İÇ HIRSIZLIK =====
    if len(internal_df) > 0:
        ws6 = wb.create_sheet("İÇ HIRSIZLIK")
        ws6['A1'] = "Satış Fiyatı ≥ 100 TL | Fark büyüdükçe risk AZALIR"
        ws6['A1'].font = subtitle_font
        
        for col, h in enumerate(internal_df.columns, 1):
            cell = ws6.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(internal_df.head(100).values, 4):
            for c_idx, val in enumerate(row, 1):
                ws6.cell(row=r_idx, column=c_idx, value=val)
    
    # ===== AİLE ANALİZİ =====
    if len(family_df) > 0:
        ws7 = wb.create_sheet("AİLE ANALİZİ")
        ws7['A1'] = "Benzer Ürün Ailesi - Kod Karışıklığı Tespiti"
        ws7['A1'].font = subtitle_font
        
        for col, h in enumerate(family_df.columns, 1):
            cell = ws7.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(family_df.head(100).values, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws7.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = wrap_alignment
    
    # ===== FİRE MANİPÜLASYONU =====
    if len(fire_manip_df) > 0:
        ws8 = wb.create_sheet("FİRE MANİPÜLASYONU")
        for col, h in enumerate(fire_manip_df.columns, 1):
            cell = ws8.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(fire_manip_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws8.cell(row=r_idx, column=c_idx, value=val)
    
    # ===== KASA AKTİVİTESİ =====
    if len(kasa_df) > 0:
        ws9 = wb.create_sheet("KASA AKTİVİTESİ")
        ws9['A1'] = "⚠️ KASA AKTİVİTESİ ÜRÜNLERİ - FAZLA (+) OLANLAR MANİPÜLASYON RİSKİ!"
        ws9['A1'].font = Font(bold=True, size=12, color='FF0000')
        
        for col, h in enumerate(kasa_df.columns, 1):
            cell = ws9.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(kasa_df.values, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws9.cell(row=r_idx, column=c_idx, value=val)
                # FAZLA olanları kırmızı yap
                if kasa_df.columns[c_idx-1] == 'TOPLAM' and val > 0:
                    cell.fill = PatternFill('solid', fgColor='FFCCCC')
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_excel_report(store_df, sm_df, bs_df, params):
    """Excel raporu oluştur"""
    wb = Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    kritik_fill = PatternFill('solid', fgColor='FF4444')
    riskli_fill = PatternFill('solid', fgColor='FF8800')
    dikkat_fill = PatternFill('solid', fgColor='FFCC00')
    temiz_fill = PatternFill('solid', fgColor='00CC66')
    title_font = Font(bold=True, size=14)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ===== BÖLGE ÖZETİ =====
    ws = wb.active
    ws.title = "BÖLGE ÖZETİ"
    
    ws['A1'] = "BÖLGE ENVANTER DASHBOARD"
    ws['A1'].font = title_font
    ws['A2'] = f"Dönem: {params.get('donem', '')} | Mağaza: {len(store_df)}"
    
    # Toplamlar
    ws['A4'] = "GENEL METRIKLER"
    ws['A4'].font = Font(bold=True, size=11)
    
    toplam_satis = store_df['Satış'].sum()
    toplam_fark = store_df['Fark'].sum()
    toplam_fire = store_df['Fire'].sum()
    
    ws['A5'] = "Toplam Satış"
    ws['B5'] = f"{toplam_satis:,.0f} TL"
    ws['A6'] = "Toplam Fark"
    ws['B6'] = f"{toplam_fark:,.0f} TL"
    ws['A7'] = "Toplam Fire"
    ws['B7'] = f"{toplam_fire:,.0f} TL"
    ws['A8'] = "Genel Toplam %"
    ws['B8'] = f"%{abs(toplam_fark)/toplam_satis*100:.2f}" if toplam_satis > 0 else "0%"
    
    # Risk dağılımı
    ws['A10'] = "RİSK DAĞILIMI"
    ws['A10'].font = Font(bold=True, size=11)
    
    kritik = len(store_df[store_df['Risk'].str.contains('KRİTİK')])
    riskli = len(store_df[store_df['Risk'].str.contains('RİSKLİ')])
    dikkat = len(store_df[store_df['Risk'].str.contains('DİKKAT')])
    temiz = len(store_df[store_df['Risk'].str.contains('TEMİZ')])
    
    ws['A11'] = "🔴 KRİTİK"
    ws['B11'] = kritik
    ws['A12'] = "🟠 RİSKLİ"
    ws['B12'] = riskli
    ws['A13'] = "🟡 DİKKAT"
    ws['B13'] = dikkat
    ws['A14'] = "🟢 TEMİZ"
    ws['B14'] = temiz
    
    # ===== SM ÖZETİ =====
    if len(sm_df) > 0:
        ws2 = wb.create_sheet("SM BAZLI")
        headers = ['Satış Müdürü', 'Mağaza', 'Satış', 'Fark', 'Toplam %', 'Sigara', 'İç Hırs.', 'Kritik', 'Ort.Risk', 'Risk']
        
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, (_, row) in enumerate(sm_df.iterrows(), start=2):
            ws2.cell(row=row_idx, column=1, value=row['SM']).border = border
            ws2.cell(row=row_idx, column=2, value=row['Mağaza Sayısı']).border = border
            ws2.cell(row=row_idx, column=3, value=f"{row['Satış']:,.0f}").border = border
            ws2.cell(row=row_idx, column=4, value=f"{row['Fark']:,.0f}").border = border
            ws2.cell(row=row_idx, column=5, value=f"%{row['Toplam %']:.2f}").border = border
            ws2.cell(row=row_idx, column=6, value=row['Sigara']).border = border
            ws2.cell(row=row_idx, column=7, value=row['İç Hırs.']).border = border
            ws2.cell(row=row_idx, column=8, value=row.get('Kritik Mağaza', 0)).border = border
            ws2.cell(row=row_idx, column=9, value=f"{row['Ort. Risk']:.0f}").border = border
            
            risk_cell = ws2.cell(row=row_idx, column=10, value=row['Risk'])
            risk_cell.border = border
            if 'KRİTİK' in row['Risk']:
                risk_cell.fill = kritik_fill
                risk_cell.font = Font(bold=True, color='FFFFFF')
            elif 'RİSKLİ' in row['Risk']:
                risk_cell.fill = riskli_fill
    
    # ===== BS ÖZETİ =====
    if len(bs_df) > 0:
        ws3 = wb.create_sheet("BS BAZLI")
        headers = ['Bölge Sorumlusu', 'Mağaza', 'Satış', 'Fark', 'Toplam %', 'Sigara', 'İç Hırs.', 'Kritik', 'Ort.Risk', 'Risk']
        
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, (_, row) in enumerate(bs_df.iterrows(), start=2):
            ws3.cell(row=row_idx, column=1, value=row['BS']).border = border
            ws3.cell(row=row_idx, column=2, value=row['Mağaza Sayısı']).border = border
            ws3.cell(row=row_idx, column=3, value=f"{row['Satış']:,.0f}").border = border
            ws3.cell(row=row_idx, column=4, value=f"{row['Fark']:,.0f}").border = border
            ws3.cell(row=row_idx, column=5, value=f"%{row['Toplam %']:.2f}").border = border
            ws3.cell(row=row_idx, column=6, value=row['Sigara']).border = border
            ws3.cell(row=row_idx, column=7, value=row['İç Hırs.']).border = border
            ws3.cell(row=row_idx, column=8, value=row.get('Kritik Mağaza', 0)).border = border
            ws3.cell(row=row_idx, column=9, value=f"{row['Ort. Risk']:.0f}").border = border
            
            risk_cell = ws3.cell(row=row_idx, column=10, value=row['Risk'])
            risk_cell.border = border
            if 'KRİTİK' in row['Risk']:
                risk_cell.fill = kritik_fill
                risk_cell.font = Font(bold=True, color='FFFFFF')
            elif 'RİSKLİ' in row['Risk']:
                risk_cell.fill = riskli_fill
    
    # ===== MAĞAZA DETAY =====
    ws4 = wb.create_sheet("MAĞAZA DETAY")
    headers = ['Kod', 'Mağaza', 'SM', 'BS', 'Satış', 'Fark', 'Toplam %', 
               'Sigara', 'İç Hırs.', 'Kronik', '10TL', 'Risk Puan', 'Risk', 'Nedenler']
    
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row_idx, (_, row) in enumerate(store_df.iterrows(), start=2):
        ws4.cell(row=row_idx, column=1, value=row['Mağaza Kodu']).border = border
        ws4.cell(row=row_idx, column=2, value=row['Mağaza Adı'][:25] if row['Mağaza Adı'] else '').border = border
        ws4.cell(row=row_idx, column=3, value=row['SM'][:15] if row['SM'] else '').border = border
        ws4.cell(row=row_idx, column=4, value=row['BS'][:15] if row['BS'] else '').border = border
        ws4.cell(row=row_idx, column=5, value=f"{row['Satış']:,.0f}").border = border
        ws4.cell(row=row_idx, column=6, value=f"{row['Fark']:,.0f}").border = border
        ws4.cell(row=row_idx, column=7, value=f"%{row['Toplam %']:.2f}").border = border
        ws4.cell(row=row_idx, column=8, value=row['Sigara']).border = border
        ws4.cell(row=row_idx, column=9, value=row['İç Hırs.']).border = border
        ws4.cell(row=row_idx, column=10, value=row['Kronik']).border = border
        ws4.cell(row=row_idx, column=11, value=f"{row['10TL Adet']:.0f}").border = border
        ws4.cell(row=row_idx, column=12, value=f"{row['Risk Puan']:.0f}").border = border
        
        risk_cell = ws4.cell(row=row_idx, column=13, value=row['Risk'])
        risk_cell.border = border
        if 'KRİTİK' in row['Risk']:
            risk_cell.fill = kritik_fill
            risk_cell.font = Font(bold=True, color='FFFFFF')
        elif 'RİSKLİ' in row['Risk']:
            risk_cell.fill = riskli_fill
            risk_cell.font = Font(bold=True, color='FFFFFF')
        elif 'DİKKAT' in row['Risk']:
            risk_cell.fill = dikkat_fill
        else:
            risk_cell.fill = temiz_fill
        
        ws4.cell(row=row_idx, column=14, value=row['Nedenler']).border = border
    
    # Sütun genişlikleri
    for ws in [ws2, ws3, ws4] if len(sm_df) > 0 else [ws4]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ========== ANA UYGULAMA ==========

st.title("🌍 Bölge Dashboard")

# Mod seçimi
data_source = st.radio("📊 Veri Kaynağı", ["📁 Excel Yükle", "☁️ Supabase"], horizontal=True)

if data_source == "☁️ Supabase":
    # Supabase'den veri çek
    col_sm, col_donem = st.columns([1, 1])
    
    available_sms = get_available_sms_from_supabase()
    available_periods = get_available_periods_from_supabase()
    
    with col_sm:
        if available_sms:
            sm_options = ["📊 TÜMÜ (Bölge)"] + available_sms
            selected_sm_option = st.selectbox("👔 Satış Müdürü", sm_options)
            
            if selected_sm_option == "📊 TÜMÜ (Bölge)":
                selected_sm = None
            else:
                selected_sm = selected_sm_option
        else:
            st.warning("Henüz veri yüklenmemiş")
            selected_sm = None
            selected_sm_option = None
    
    with col_donem:
        if available_periods:
            selected_periods = st.multiselect("📅 Dönem", available_periods, default=available_periods[:1])
        else:
            selected_periods = []
    
    if selected_sm_option and selected_periods:
        with st.spinner("Veriler yükleniyor..."):
            df_raw = get_data_from_supabase(satis_muduru=selected_sm, donemler=selected_periods)
        
        if len(df_raw) == 0:
            st.warning("Seçilen kriterlere uygun veri bulunamadı")
        else:
            st.success(f"✅ {len(df_raw):,} satır yüklendi")
            
            df = analyze_inventory(df_raw)
            
            params = {
                'donem': ', '.join(selected_periods),
                'tarih': datetime.now().strftime('%Y-%m-%d'),
            }
            
            # Analiz
            with st.spinner("🔄 Analiz ediliyor..."):
                store_df, bolge_ort = analyze_all_stores(df)
                sm_df = aggregate_by_group(store_df, 'SM')
                bs_df = aggregate_by_group(store_df, 'BS')
            
            if len(store_df) == 0:
                st.error("Analiz edilecek mağaza bulunamadı!")
            else:
                # Bölge toplamları
                toplam_satis = store_df['Satış'].sum()
                toplam_fark = store_df['Fark'].sum()
                toplam_fire = store_df['Fire'].sum()
                toplam_acik = store_df['Toplam Açık'].sum()
                toplam_gun = store_df['Gün'].sum()
                
                # Oranlar
                fark_oran = abs(toplam_fark) / toplam_satis * 100 if toplam_satis > 0 else 0
                fire_oran = abs(toplam_fire) / toplam_satis * 100 if toplam_satis > 0 else 0
                toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
                gunluk_fark = toplam_fark / toplam_gun if toplam_gun > 0 else 0
                gunluk_fire = toplam_fire / toplam_gun if toplam_gun > 0 else 0
                
                # Risk sayıları
                kritik_sayisi = len(store_df[store_df['Risk'].str.contains('KRİTİK')])
                riskli_sayisi = len(store_df[store_df['Risk'].str.contains('RİSKLİ')])
                dikkat_sayisi = len(store_df[store_df['Risk'].str.contains('DİKKAT')])
                temiz_sayisi = len(store_df[store_df['Risk'].str.contains('TEMİZ')])
                
                # 10TL Özet
                toplam_10tl_adet = store_df['10TL Adet'].sum()
                toplam_10tl_tutar = store_df['10TL Tutar'].sum()
                
                # ========== GÖRÜNÜM ==========
                st.markdown("---")
                display_sm_name = selected_sm if selected_sm else "Bölge"
                st.subheader(f"📊 {display_sm_name} - {len(store_df)} Mağaza")
                
                # Üst metrikler
                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("💰 Satış", f"{toplam_satis/1e6:.1f}M TL")
                col2.metric("📉 Fark", f"%{fark_oran:.2f}", f"{toplam_fark/1000:.0f}K | Gün: {gunluk_fark/1000:.1f}K")
                col3.metric("🔥 Fire", f"%{fire_oran:.2f}", f"{toplam_fire/1000:.0f}K | Gün: {gunluk_fire/1000:.1f}K")
                col4.metric("📊 Toplam", f"%{toplam_oran:.2f}", f"{toplam_acik/1000:.0f}K")
                
                if toplam_10tl_adet != 0:
                    col5.metric("💰 10 TL", f"{toplam_10tl_adet:+.0f} / {toplam_10tl_tutar:,.0f}₺", 
                               "FAZLA" if toplam_10tl_adet > 0 else "AÇIK")
                else:
                    col5.metric("💰 10 TL", "0", "TAMAM")
                
                # Risk dağılımı
                st.markdown("### 📊 Risk Dağılımı")
                r1, r2, r3, r4 = st.columns(4)
                r1.markdown(f'<div class="risk-kritik">🔴 KRİTİK: {kritik_sayisi}</div>', unsafe_allow_html=True)
                r2.markdown(f'<div class="risk-riskli">🟠 RİSKLİ: {riskli_sayisi}</div>', unsafe_allow_html=True)
                r3.markdown(f'<div class="risk-dikkat">🟡 DİKKAT: {dikkat_sayisi}</div>', unsafe_allow_html=True)
                r4.markdown(f'<div class="risk-temiz">🟢 TEMİZ: {temiz_sayisi}</div>', unsafe_allow_html=True)
                
                # Sekmeler
                tabs = st.tabs(["👔 SM Özet", "📋 BS Özet", "🏪 Mağazalar", "📊 Top 10", "📥 İndir"])
                
                with tabs[0]:
                    st.subheader("👔 Satış Müdürü Bazlı Özet")
                    if len(sm_df) > 0:
                        for _, row in sm_df.iterrows():
                            with st.expander(f"**{row['SM']}** - {row['Mağaza Sayısı']} Mağaza | Kayıp: %{row['Toplam %']:.1f}"):
                                c1, c2, c3, c4, c5 = st.columns(5)
                                c1.metric("Satış", f"{row['Satış']/1e6:.1f}M")
                                c2.metric("Fark", f"{row['Fark']/1000:.0f}K", f"%{row['Fark %']:.1f}")
                                c3.metric("Fire", f"{row['Fire']/1000:.0f}K", f"%{row['Fire %']:.1f}")
                                c4.metric("🚬 Sigara", f"{row['Sigara']:.0f}")
                                c5.metric("🔒 İç Hırs.", f"{row['İç Hırs.']:.0f}")
                    else:
                        st.info("SM verisi bulunamadı")
                
                with tabs[1]:
                    st.subheader("📋 Bölge Sorumlusu Bazlı Özet")
                    if len(bs_df) > 0:
                        for _, row in bs_df.iterrows():
                            col1, col2, col3, col4, col5, col6 = st.columns([2, 1, 1, 1, 1, 1])
                            col1.write(f"**{row['BS']}** ({row['Mağaza Sayısı']} mağ.)")
                            col2.write(f"Satış: {row['Satış']/1e6:.1f}M")
                            col3.write(f"Fark: {row['Fark']/1000:.0f}K")
                            col4.write(f"Fire: {row['Fire']/1000:.0f}K")
                            col5.write(f"Kayıp: %{row['Toplam %']:.1f}")
                            col6.write(f"🚬{row['Sigara']:.0f} 🔒{row['İç Hırs.']:.0f}")
                    else:
                        st.info("BS verisi bulunamadı")
                
                with tabs[2]:
                    st.subheader("🏪 Tüm Mağazalar")
                    display_cols = ['Mağaza Kodu', 'Mağaza Adı', 'SM', 'BS', 'Satış', 'Fark', 'Fark %', 
                                   'Fire', 'Fire %', 'Toplam Açık', 'Toplam %', 'Sigara', 'İç Hırs.', 'Risk Puan', 'Risk']
                    display_cols = [c for c in display_cols if c in store_df.columns]
                    st.dataframe(store_df[display_cols].sort_values('Risk Puan', ascending=False), 
                                use_container_width=True, height=500)
                
                with tabs[3]:
                    st.subheader("📊 En Riskli 10 Mağaza")
                    top10 = store_df.nlargest(10, 'Risk Puan')
                    for _, row in top10.iterrows():
                        risk_text = row.get('Risk', '')
                        if 'KRİTİK' in str(risk_text):
                            st.error(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}** | Risk: {row['Risk Puan']:.0f}\n\n"
                                    f"Fark: {row['Fark']/1000:.0f}K | Fire: {row['Fire']/1000:.0f}K | Kayıp: %{row['Toplam %']:.1f}\n\n"
                                    f"🚬 Sigara: {row['Sigara']:.0f} | 🔒 İç Hırs: {row['İç Hırs.']:.0f}")
                        elif 'RİSKLİ' in str(risk_text):
                            st.warning(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}** | Risk: {row['Risk Puan']:.0f}\n\n"
                                      f"Fark: {row['Fark']/1000:.0f}K | Fire: {row['Fire']/1000:.0f}K | Kayıp: %{row['Toplam %']:.1f}\n\n"
                                      f"🚬 Sigara: {row['Sigara']:.0f} | 🔒 İç Hırs: {row['İç Hırs.']:.0f}")
                        else:
                            st.info(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}** | Risk: {row['Risk Puan']:.0f}\n\n"
                                   f"Fark: {row['Fark']/1000:.0f}K | Fire: {row['Fire']/1000:.0f}K | Kayıp: %{row['Toplam %']:.1f}")
                
                with tabs[4]:
                    st.subheader("📥 Rapor İndir")
                    
                    excel_data = create_dashboard_excel(store_df, sm_df, bs_df, params)
                    
                    st.download_button(
                        label="📥 Bölge Dashboard Excel",
                        data=excel_data,
                        file_name=f"BOLGE_DASHBOARD_{params.get('donem', '')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

else:
    # Excel yükleme modu
    uploaded_file = st.file_uploader("📁 Envanter Excel Yükle", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        try:
            # Dosyayı oku
            xl = pd.ExcelFile(uploaded_file)
            sheet_names = xl.sheet_names
            
            best_sheet = None
            max_cols = 0
            for sheet in sheet_names:
                temp_df = pd.read_excel(uploaded_file, sheet_name=sheet, nrows=5)
                if len(temp_df.columns) > max_cols:
                    max_cols = len(temp_df.columns)
                    best_sheet = sheet
            
            df_raw = pd.read_excel(uploaded_file, sheet_name=best_sheet)
            st.success(f"✅ {len(df_raw):,} satır | {len(df_raw.columns)} sütun")
            
            df = analyze_inventory(df_raw)
            
            params = {
                'donem': str(df['Envanter Dönemi'].iloc[0]) if 'Envanter Dönemi' in df.columns else '',
                'tarih': str(df['Envanter Tarihi'].iloc[0])[:10] if 'Envanter Tarihi' in df.columns else '',
            }
            
            # Analiz
            with st.spinner("🔄 Analiz ediliyor..."):
                store_df, bolge_ort = analyze_all_stores(df)
                sm_df = aggregate_by_group(store_df, 'SM')
                bs_df = aggregate_by_group(store_df, 'BS')
            
            if len(store_df) == 0:
                st.error("Analiz edilecek mağaza bulunamadı!")
            else:
                # Bölge toplamları
                toplam_satis = store_df['Satış'].sum()
                toplam_fark = store_df['Fark'].sum()  # Fark + Kısmi
                toplam_fire = store_df['Fire'].sum()
                toplam_acik = store_df['Toplam Açık'].sum()  # Fark + Kısmi + Fire
                toplam_gun = store_df['Gün'].sum()
                
                # Oranlar
                fark_oran = abs(toplam_fark) / toplam_satis * 100 if toplam_satis > 0 else 0
                fire_oran = abs(toplam_fire) / toplam_satis * 100 if toplam_satis > 0 else 0
                toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
                gunluk_fark = toplam_fark / toplam_gun if toplam_gun > 0 else 0
            gunluk_fire = toplam_fire / toplam_gun if toplam_gun > 0 else 0
            
            # Risk sayıları
            kritik = len(store_df[store_df['Risk'].str.contains('KRİTİK')])
            riskli = len(store_df[store_df['Risk'].str.contains('RİSKLİ')])
            dikkat = len(store_df[store_df['Risk'].str.contains('DİKKAT')])
            temiz = len(store_df[store_df['Risk'].str.contains('TEMİZ')])
            
            # ===== ÜST METRİKLER =====
            st.markdown(f"### 📊 Dönem: {params['donem']} | {len(store_df)} Mağaza")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("💰 Toplam Satış", f"{toplam_satis/1_000_000:.1f}M TL")
            with col2:
                st.metric("📉 Fark", f"{toplam_fark:,.0f} TL", f"%{fark_oran:.2f} | Günlük: {gunluk_fark:,.0f}₺")
            with col3:
                st.metric("🔥 Fire", f"{toplam_fire:,.0f} TL", f"%{fire_oran:.2f} | Günlük: {gunluk_fire:,.0f}₺")
            with col4:
                st.metric("📊 Toplam", f"{toplam_acik:,.0f} TL", f"%{toplam_oran:.2f}")
            
            # Risk dağılımı
            st.markdown("### 📊 Risk Dağılımı")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.markdown(f'<div class="risk-kritik">🔴 KRİTİK<br>{kritik}</div>', unsafe_allow_html=True)
            with col2:
                st.markdown(f'<div class="risk-riskli">🟠 RİSKLİ<br>{riskli}</div>', unsafe_allow_html=True)
            with col3:
                st.markdown(f'<div class="risk-dikkat">🟡 DİKKAT<br>{dikkat}</div>', unsafe_allow_html=True)
            with col4:
                st.markdown(f'<div class="risk-temiz">🟢 TEMİZ<br>{temiz}</div>', unsafe_allow_html=True)
            
            # ===== SEKMELER =====
            tabs = st.tabs(["🏆 Top 10", "👔 SM Bazlı", "👤 BS Bazlı", "🏪 Tüm Mağazalar", "📥 İndir"])
            
            # TOP 10
            with tabs[0]:
                st.markdown("### 🚨 En Riskli 10 Mağaza")
                top10 = store_df.head(10)
                
                for idx, (_, row) in enumerate(top10.iterrows()):
                    risk_class = row['Risk Class']
                    col1, col2, col3 = st.columns([1, 3, 0.5])
                    
                    with col1:
                        st.markdown(f"""
                        <div style="background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); 
                                    border-radius: 10px; padding: 15px; color: white;
                                    border-left: 5px solid {'#ff4444' if risk_class=='kritik' else '#ff8800' if risk_class=='riskli' else '#ffcc00' if risk_class=='dikkat' else '#00cc66'};">
                            <h3 style="margin:0; color: white;">{row['Mağaza Kodu']}</h3>
                            <p style="margin:5px 0; font-size: 0.9rem;">{row['Mağaza Adı'][:20] if row['Mağaza Adı'] else ''}</p>
                            <h2 style="margin:10px 0; color: {'#ff4444' if risk_class=='kritik' else '#ff8800' if risk_class=='riskli' else '#ffcc00'};">
                                Risk: {row['Risk Puan']:.0f}
                            </h2>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col2:
                        c1, c2, c3, c4, c5 = st.columns(5)
                        c1.metric("🚬 Sigara", row['Sigara'])
                        c2.metric("🔒 İç Hırs.", row['İç Hırs.'])
                        c3.metric("📉 Fark", f"{row['Fark']:,.0f}", f"%{row['Fark %']:.1f}")
                        c4.metric("🔥 Fire", f"{row['Fire']:,.0f}", f"%{row['Fire %']:.1f}")
                        c5.metric("📊 Toplam", f"{row['Toplam Açık']:,.0f}", f"%{row['Toplam %']:.1f}")
                        
                        if row['Nedenler'] != "-":
                            st.caption(f"**Nedenler:** {row['Nedenler']}")
                    
                    with col3:
                        # İndirme butonu - detaylı rapor
                        mag_adi_clean = row['Mağaza Adı'].replace(' ', '_').replace('/', '_')[:30] if row['Mağaza Adı'] else ''
                        report_data = create_store_report(row, params, df)
                        st.download_button(
                            label="📥",
                            data=report_data,
                            file_name=f"{row['Mağaza Kodu']}_{mag_adi_clean}_Risk_Raporu.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"top10_dl_{idx}"
                        )
                    
                    st.divider()
            
            # SM BAZLI
            with tabs[1]:
                st.markdown("### 👔 Satış Müdürleri Karşılaştırma")
                if len(sm_df) > 0:
                    display_cols = ['SM', 'Mağaza Sayısı', 'Satış', 'Fark', 'Günlük Fark', 'Fire', 'Günlük Fire', 'Toplam %', 'Fire %', 'Sigara', 'Kritik Mağaza', 'Ort. Risk', 'Risk']
                    display_sm = sm_df[display_cols].copy()
                    display_sm['Satış'] = display_sm['Satış'].apply(lambda x: f"{x/1_000_000:.1f}M")
                    display_sm['Fark'] = display_sm['Fark'].apply(lambda x: f"{x:,.0f}")
                    display_sm['Günlük Fark'] = display_sm['Günlük Fark'].apply(lambda x: f"{x:,.0f}")
                    display_sm['Fire'] = display_sm['Fire'].apply(lambda x: f"{x:,.0f}")
                    display_sm['Günlük Fire'] = display_sm['Günlük Fire'].apply(lambda x: f"{x:,.0f}")
                    display_sm['Toplam %'] = display_sm['Toplam %'].apply(lambda x: f"%{x:.2f}")
                    display_sm['Fire %'] = display_sm['Fire %'].apply(lambda x: f"%{x:.2f}")
                    display_sm['Ort. Risk'] = display_sm['Ort. Risk'].apply(lambda x: f"{x:.0f}")
                    st.dataframe(display_sm, use_container_width=True, hide_index=True)
                    
                    # SM Detay
                    st.markdown("---")
                    selected_sm = st.selectbox("📋 SM Detay Göster", sm_df['SM'].tolist())
                    if selected_sm:
                        sm_row = sm_df[sm_df['SM'] == selected_sm].iloc[0]
                        sm_magazalar = store_df[store_df['SM'] == selected_sm]
                        
                        # SM Özet metrikleri
                        st.markdown(f"#### {selected_sm} - Özet")
                        c1, c2, c3, c4 = st.columns(4)
                        c1.metric("📊 Mağaza", f"{len(sm_magazalar)}")
                        c2.metric("📉 Fark", f"{sm_row['Fark']:,.0f}₺", f"Günlük: {sm_row['Günlük Fark']:,.0f}₺")
                        c3.metric("🔥 Fire", f"{sm_row['Fire']:,.0f}₺", f"Günlük: {sm_row['Günlük Fire']:,.0f}₺")
                        c4.metric("📊 Risk", f"{sm_row['Ort. Risk']:.0f}")
                        
                        # BS'ler
                        st.markdown("##### 👤 Bölge Sorumluları")
                        sm_bs_list = sm_magazalar['BS'].unique().tolist()
                        for bs_name in sm_bs_list:
                            bs_mag = sm_magazalar[sm_magazalar['BS'] == bs_name]
                            bs_fark = bs_mag['Fark'].sum()
                            bs_fire = bs_mag['Fire'].sum()
                            bs_risk = bs_mag['Risk Puan'].mean()
                            bs_sigara = bs_mag['Sigara'].sum()
                            st.info(f"**{bs_name}**: {len(bs_mag)} mağaza | Fark: {bs_fark:,.0f}₺ | Fire: {bs_fire:,.0f}₺ | Risk: {bs_risk:.0f} | 🚬 {bs_sigara}")
                        
                        # Mağaza listesi
                        st.markdown("##### 🏪 Mağazalar")
                        show_cols = ['Mağaza Kodu', 'Mağaza Adı', 'BS', 'Fark', 'Günlük Fark', 'Toplam %', 'Sigara', 'İç Hırs.', 'Risk Puan', 'Risk']
                        st.dataframe(sm_magazalar[show_cols], use_container_width=True, hide_index=True)
                else:
                    st.info("SM verisi bulunamadı")
            
            # BS BAZLI
            with tabs[2]:
                st.markdown("### 👤 Bölge Sorumluları Karşılaştırma")
                if len(bs_df) > 0:
                    display_cols = ['BS', 'Mağaza Sayısı', 'Satış', 'Fark', 'Günlük Fark', 'Fire', 'Günlük Fire', 'Toplam %', 'Fire %', 'Sigara', 'Kritik Mağaza', 'Ort. Risk', 'Risk']
                    display_bs = bs_df[display_cols].copy()
                    display_bs['Satış'] = display_bs['Satış'].apply(lambda x: f"{x/1_000_000:.1f}M")
                    display_bs['Fark'] = display_bs['Fark'].apply(lambda x: f"{x:,.0f}")
                    display_bs['Günlük Fark'] = display_bs['Günlük Fark'].apply(lambda x: f"{x:,.0f}")
                    display_bs['Fire'] = display_bs['Fire'].apply(lambda x: f"{x:,.0f}")
                    display_bs['Günlük Fire'] = display_bs['Günlük Fire'].apply(lambda x: f"{x:,.0f}")
                    display_bs['Toplam %'] = display_bs['Toplam %'].apply(lambda x: f"%{x:.2f}")
                    display_bs['Fire %'] = display_bs['Fire %'].apply(lambda x: f"%{x:.2f}")
                    display_bs['Ort. Risk'] = display_bs['Ort. Risk'].apply(lambda x: f"{x:.0f}")
                    st.dataframe(display_bs, use_container_width=True, hide_index=True)
                    
                    # BS Detay
                    st.markdown("---")
                    selected_bs = st.selectbox("📋 BS Detay Göster", bs_df['BS'].tolist())
                    if selected_bs:
                        bs_row = bs_df[bs_df['BS'] == selected_bs].iloc[0]
                        bs_magazalar = store_df[store_df['BS'] == selected_bs]
                        
                        # BS Özet metrikleri
                        st.markdown(f"#### {selected_bs} - Özet")
                        c1, c2, c3, c4 = st.columns(4)
                        c1.metric("📊 Mağaza", f"{len(bs_magazalar)}")
                        c2.metric("📉 Fark", f"{bs_row['Fark']:,.0f}₺", f"Günlük: {bs_row['Günlük Fark']:,.0f}₺")
                        c3.metric("🔥 Fire", f"{bs_row['Fire']:,.0f}₺", f"Günlük: {bs_row['Günlük Fire']:,.0f}₺")
                        c4.metric("📊 Risk", f"{bs_row['Ort. Risk']:.0f}")
                        
                        # Mağaza listesi indirme butonlu
                        st.markdown("##### 🏪 Mağazalar")
                        for idx, (_, row) in enumerate(bs_magazalar.iterrows()):
                            col1, col2 = st.columns([5, 1])
                            with col1:
                                sigara_txt = f"🚬 {row['Sigara']}" if row['Sigara'] > 0 else ""
                                st.write(f"**{row['Mağaza Kodu']}** - {row['Mağaza Adı'][:25]} | Fark: {row['Fark']:,.0f}₺ | Risk: {row['Risk Puan']:.0f} {sigara_txt}")
                            with col2:
                                mag_adi_clean = row['Mağaza Adı'].replace(' ', '_').replace('/', '_')[:30] if row['Mağaza Adı'] else ''
                                report_data = create_store_report(row, params, df)
                                st.download_button("📥", data=report_data, file_name=f"{row['Mağaza Kodu']}_{mag_adi_clean}_Risk_Raporu.xlsx", 
                                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"bs_dl_{idx}")
                else:
                    st.info("BS verisi bulunamadı")
            
            # TÜM MAĞAZALAR
            with tabs[3]:
                st.markdown("### 🏪 Tüm Mağazalar")
                
                # Filtreler
                col1, col2, col3 = st.columns(3)
                with col1:
                    risk_filter = st.multiselect("Risk Filtre", ["🔴 KRİTİK", "🟠 RİSKLİ", "🟡 DİKKAT", "🟢 TEMİZ"])
                with col2:
                    sm_filter = st.multiselect("SM Filtre", store_df['SM'].unique().tolist())
                with col3:
                    bs_filter = st.multiselect("BS Filtre", store_df['BS'].unique().tolist())
                
                filtered_df = store_df.copy()
                if risk_filter:
                    filtered_df = filtered_df[filtered_df['Risk'].isin(risk_filter)]
                if sm_filter:
                    filtered_df = filtered_df[filtered_df['SM'].isin(sm_filter)]
                if bs_filter:
                    filtered_df = filtered_df[filtered_df['BS'].isin(bs_filter)]
                
                st.info(f"📊 {len(filtered_df)} mağaza gösteriliyor")
                
                show_cols = ['Mağaza Kodu', 'Mağaza Adı', 'SM', 'BS', 'Satış', 'Fark', 'Toplam %', 
                            'Sigara', 'İç Hırs.', '10TL Adet', '10TL Tutar', 'Risk Puan', 'Risk']
                display_filtered = filtered_df[show_cols].copy()
                display_filtered['Satış'] = display_filtered['Satış'].apply(lambda x: f"{x:,.0f}")
                display_filtered['Fark'] = display_filtered['Fark'].apply(lambda x: f"{x:,.0f}")
                display_filtered['Toplam %'] = display_filtered['Toplam %'].apply(lambda x: f"%{x:.1f}")
                display_filtered['10TL Tutar'] = display_filtered['10TL Tutar'].apply(lambda x: f"{x:,.0f}")
                display_filtered['Risk Puan'] = display_filtered['Risk Puan'].apply(lambda x: f"{x:.0f}")
                
                st.dataframe(display_filtered, use_container_width=True, hide_index=True)
            
            # İNDİR
            with tabs[4]:
                st.markdown("### 📥 Rapor İndir")
                
                excel_data = create_excel_report(store_df, sm_df, bs_df, params)
                
                st.download_button(
                    label="📥 Bölge Dashboard Excel",
                    data=excel_data,
                    file_name=f"BOLGE_DASHBOARD_{params['donem']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.markdown("---")
                st.markdown("""
                **Excel İçeriği:**
                - 📋 Bölge Özeti
                - 👔 SM Bazlı Analiz
                - 👤 BS Bazlı Analiz  
                - 🏪 Mağaza Detay (Risk puanına göre sıralı)
                """)
        
        except Exception as e:
            st.error(f"Hata: {str(e)}")
            st.exception(e)
    
    else:
        st.info("👆 Envanter Excel dosyası yükleyin")
        
        st.markdown("""
        ### 📊 Dashboard Özellikleri
        
        **Hiyerarşik Görünüm:**
        - 🌍 Bölge Toplamları
        - 👔 SM (Satış Müdürü) Bazlı
        - 👤 BS (Bölge Sorumlusu) Bazlı
        - 🏪 Mağaza Bazlı
        
        **Risk Skorlama (0-100):**
        | Kriter | Ağırlık |
        |--------|---------|
        | Kayıp Oranı | %30 |
        | Sigara Açığı | %30 |
        | İç Hırsızlık | %30 |
        | Kronik Açık | %5 |
        | 10TL Ürünleri | %5 |
        
        **Karşılaştırma:** Bölge ortalamasına göre
        """)
